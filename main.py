"""
LeadHarvest — Main entry point.

Run from the terminal:
    python main.py

You will be prompted to select a business category and Nigerian city.
The scraper will then:
  1. Search Google Places for matching businesses
  2. Visit each website to extract contacts and score quality
  3. Export results to a timestamped Excel file

Output is saved to output/exports/
"""

import asyncio
import os
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from db.database import filter_new_businesses, get_all_businesses, get_campaign_status_map, get_existing_website_urls, save_businesses
from models.business import Business
from scraper.places import CATEGORY_MAP, search_businesses
from scraper.website import scrape_all_websites
from utils.logger import get_logger

load_dotenv()
logger = get_logger(__name__)

EXPORT_PATH = os.getenv("EXPORT_PATH", "output/exports/")


# ── Excel export ───────────────────────────────────────────────────────────────

def _style_header_row(ws, row: int = 1) -> None:
    """Applies bold, coloured header styling to the first row of a worksheet."""
    header_fill = PatternFill("solid", fgColor="2E5FA3")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for cell in ws[row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _auto_fit_columns(ws) -> None:
    """Sets reasonable column widths based on content."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                cell_len = len(str(cell.value)) if cell.value else 0
                max_len = max(max_len, cell_len)
            except Exception:
                pass
        # Cap width between 12 and 60 characters
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 60)


MASTER_FILE = "leadharvest_master.xlsx"


def update_master_excel(businesses: list[Business]) -> str:
    """
    Appends new business results to the master Excel file.

    The master file contains a single 'All Results' sheet that grows with
    every scrape run. If the file does not exist yet it is created fresh.
    Dedup is performed on place_id (primary) and email (fallback) so the
    master file never accumulates duplicates even if the DB is reset.

    Args:
        businesses: New businesses from the current scrape run.

    Returns:
        Path to the master Excel file.
    """
    export_dir = Path(EXPORT_PATH)
    export_dir.mkdir(parents=True, exist_ok=True)
    master_path = export_dir / MASTER_FILE

    sent_map = get_campaign_status_map()
    new_records = [b.to_dict() for b in businesses]
    df_new = pd.DataFrame(new_records)

    if master_path.exists():
        df_existing = pd.read_excel(master_path, sheet_name="All Results")

        # Dedup: drop rows from df_new that are already in the master file.
        # Check place_id first (most reliable), then fall back to email.
        existing_place_ids = set(df_existing["Place ID"].dropna().astype(str)) if "Place ID" in df_existing.columns else set()
        existing_emails = set(df_existing["Email"].dropna().astype(str).str.lower()) if "Email" in df_existing.columns else set()

        def _is_duplicate(row: dict) -> bool:
            pid = str(row.get("Place ID") or "").strip()
            email = str(row.get("Email") or "").strip().lower()
            if pid and pid in existing_place_ids:
                return True
            if email and email in existing_emails:
                return True
            return False

        df_new = df_new[~df_new.apply(_is_duplicate, axis=1)]
        df_master = pd.concat([df_existing, df_new], ignore_index=True)
    else:
        df_master = df_new

    email_col = df_master.get("Email", pd.Series(dtype=str))
    df_master["Campaign Status"] = email_col.str.lower().map(
        lambda e: "Sent" if e in sent_map else "Not Contacted"
    )
    df_master["Date Sent"] = email_col.str.lower().map(
        lambda e: sent_map.get(e, "")
    )

    with pd.ExcelWriter(master_path, engine="openpyxl") as writer:
        df_master.to_excel(writer, sheet_name="All Results", index=False)

    # Apply styling
    wb = load_workbook(master_path)
    ws = wb["All Results"]
    _style_header_row(ws)
    _auto_fit_columns(ws)
    ws.freeze_panes = "A2"
    wb.save(master_path)

    return str(master_path)


def refresh_master_excel_campaign_status() -> None:
    """
    Re-reads the master Excel file and updates the Campaign Status and
    Date Sent columns from the current campaigns table. Called from
    Streamlit after each successful email send.
    """
    export_dir = Path(EXPORT_PATH)
    master_path = export_dir / MASTER_FILE

    if not master_path.exists():
        return

    sent_map = get_campaign_status_map()
    try:
        wb = load_workbook(master_path)
    except Exception:
        logger.warning("Master Excel file is corrupted and cannot be updated. Delete it and re-export.")
        return
    ws = wb["All Results"]

    headers = [cell.value for cell in ws[1]]

    # Find or create the two columns
    if "Campaign Status" not in headers:
        status_col = len(headers) + 1
        ws.cell(row=1, column=status_col, value="Campaign Status")
        date_col = status_col + 1
        ws.cell(row=1, column=date_col, value="Date Sent")
    else:
        status_col = headers.index("Campaign Status") + 1
        if "Date Sent" in headers:
            date_col = headers.index("Date Sent") + 1
        else:
            date_col = status_col + 1
            ws.cell(row=1, column=date_col, value="Date Sent")

    # Find email column index
    if "Email" not in headers:
        wb.save(master_path)
        return
    email_col_idx = headers.index("Email") + 1

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        email_cell = row[email_col_idx - 1]
        email = str(email_cell.value or "").strip().lower()
        if email in sent_map:
            ws.cell(row=email_cell.row, column=status_col, value="Sent")
            ws.cell(row=email_cell.row, column=date_col, value=sent_map[email])
        else:
            ws.cell(row=email_cell.row, column=status_col, value="Not Contacted")
            ws.cell(row=email_cell.row, column=date_col, value="")

    _style_header_row(ws)
    _auto_fit_columns(ws)
    wb.save(master_path)


def export_to_excel(businesses: list[Business], category: str, city: str) -> str:
    """
    Exports scrape results to a formatted Excel file with three sheets:
      - All Results
      - High Priority Leads (score < 50)
      - Summary

    Args:
        businesses: List of scraped Business objects.
        category: The business category searched.
        city: The city searched.

    Returns:
        Path to the saved Excel file.
    """
    export_dir = Path(EXPORT_PATH)
    export_dir.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_cat = category.replace(" ", "_").replace("/", "_")
    safe_city = city.replace(" ", "_")
    filename = f"leadharvest_{safe_cat}_{safe_city}_{timestamp}.xlsx"
    filepath = export_dir / filename

    # Convert all businesses to dicts
    all_records = [b.to_dict() for b in businesses]
    priority_records = [b.to_dict() for b in businesses if b.is_high_priority()]

    df_all = pd.DataFrame(all_records)
    df_priority = pd.DataFrame(priority_records) if priority_records else pd.DataFrame(columns=df_all.columns)

    # Build summary stats
    total = len(businesses)
    with_email = sum(1 for b in businesses if b.email)
    email_from_social = sum(1 for b in businesses if b.email_source and b.email_source != "website")
    with_whatsapp = sum(1 for b in businesses if b.whatsapp)
    with_website = sum(1 for b in businesses if b.website_url)
    without_website = sum(1 for b in businesses if not b.has_website)
    high_priority_count = len(priority_records)
    avg_score = (
        sum(b.website_quality_score for b in businesses if b.has_website) / with_website
        if with_website > 0 else 0
    )

    summary_data = {
        "Metric": [
            "Total Businesses Found",
            "Businesses with Website",
            "Businesses without Website (High Priority)",
            "Businesses with Email",
            "Emails Found via Social Media",
            "Businesses with WhatsApp",
            "High Priority Leads (No website OR score < 50)",
            "Average Website Quality Score",
            "Category",
            "City",
            "Scraped At",
        ],
        "Value": [
            total,
            with_website,
            without_website,
            with_email,
            email_from_social,
            with_whatsapp,
            high_priority_count,
            f"{avg_score:.1f} / 100",
            category,
            city,
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        ],
    }
    df_summary = pd.DataFrame(summary_data)

    # Write to Excel
    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        df_all.to_excel(writer, sheet_name="All Results", index=False)
        df_priority.to_excel(writer, sheet_name="High Priority Leads", index=False)
        df_summary.to_excel(writer, sheet_name="Summary", index=False)

    # Apply styling
    wb = load_workbook(filepath)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        _style_header_row(ws)
        _auto_fit_columns(ws)
        ws.freeze_panes = "A2"

    wb.save(filepath)

    return str(filepath)


def rebuild_master_excel_from_db() -> str:
    """
    Regenerates the master Excel file entirely from the SQLite database.
    Use this to recover the master file after it has been deleted or corrupted.
    Returns the path to the saved file.
    """
    export_dir = Path(EXPORT_PATH)
    export_dir.mkdir(parents=True, exist_ok=True)
    master_path = export_dir / MASTER_FILE

    records = get_all_businesses()
    sent_map = get_campaign_status_map()

    df = pd.DataFrame(records)
    if not df.empty:
        email_col = df.get("Email", pd.Series(dtype=str))
        df["Campaign Status"] = email_col.str.lower().map(
            lambda e: "Sent" if e in sent_map else "Not Contacted"
        )
        df["Date Sent"] = email_col.str.lower().map(
            lambda e: sent_map.get(e, "")
        )

    with pd.ExcelWriter(master_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="All Results", index=False)

    wb = load_workbook(master_path)
    ws = wb["All Results"]
    _style_header_row(ws)
    _auto_fit_columns(ws)
    ws.freeze_panes = "A2"
    wb.save(master_path)

    logger.info("Master Excel rebuilt from DB | rows=%d | path=%s", len(df), master_path)
    return str(master_path)


# ── CLI helpers ────────────────────────────────────────────────────────────────

def _prompt_category() -> str:
    """
    Displays preset categories for quick selection, but also accepts any
    freeform text as a custom category keyword.
    """
    categories = list(CATEGORY_MAP.keys())
    print("\nPreset business categories:")
    for i, cat in enumerate(categories, start=1):
        print(f"  [{i:>2}] {cat}")
    print()
    print("  Enter a number to use a preset, or type any custom category")
    print("  (e.g. pharmacy, gym, car dealer, supermarket)")
    print()

    while True:
        try:
            choice = input("Category (number or custom text): ").strip()
            if not choice:
                print("  Please enter a number or a category name.")
                continue
            # If the input is a number, map to the preset
            try:
                index = int(choice) - 1
                if 0 <= index < len(categories):
                    return categories[index]
                print(f"  Please enter a number between 1 and {len(categories)}, or type a custom category.")
            except ValueError:
                # Not a number — use as freeform keyword
                return choice
        except KeyboardInterrupt:
            print("\nExiting.")
            sys.exit(0)


def _prompt_city() -> str:
    """Prompts for a Nigerian city name."""
    common_cities = [
        "Lagos", "Abuja", "Port Harcourt", "Kano", "Ibadan",
        "Benin City", "Enugu", "Kaduna", "Jos", "Warri",
    ]
    print("\nCommon Nigerian cities:")
    print("  " + ", ".join(common_cities))
    print()

    while True:
        try:
            city = input("Enter city name: ").strip()
            if city:
                return city
            print("  City name cannot be empty.")
        except KeyboardInterrupt:
            print("\nExiting.")
            sys.exit(0)


def _print_banner() -> None:
    print()
    print("=" * 60)
    print("        LeadHarvest - Business Contact Scraper")
    print("        Powered by Google Places + Playwright")
    print("=" * 60)


def _print_results_preview(businesses: list[Business]) -> None:
    """Prints a quick preview of the top 5 results in the terminal."""
    print("\n--- Top 5 Results Preview ---\n")
    for i, b in enumerate(businesses[:5], start=1):
        score_label = "HIGH PRIORITY" if b.is_high_priority() else "standard"
        print(f"[{i}] {b.business_name}  [{score_label}]")
        print(f"     Phone    : {b.phone or 'N/A'}")
        print(f"     Email    : {b.email or 'N/A'}")
        print(f"     WhatsApp : {b.whatsapp or 'N/A'}")
        print(f"     Website  : {b.website_url or 'N/A'}")
        print(f"     Score    : {b.website_quality_score}/100")
        if b.website_issues:
            print(f"     Issues   : {', '.join(b.website_issues)}")
        print()


# ── Entry point ────────────────────────────────────────────────────────────────

def main() -> None:
    _print_banner()

    category = _prompt_category()
    city = _prompt_city()

    print(f"\n[INFO] Starting scrape: {category} in {city}")
    logger.info("Scrape started | category=%s | city=%s", category, city)

    # Phase 1: Fetch from Google Places (skipping URLs already in DB)
    print("\n" + "-" * 40)
    print("  STEP 1 of 2: Google Places Search")
    print("-" * 40)
    known_urls = get_existing_website_urls()
    businesses = search_businesses(category=category, city=city, known_urls=known_urls)

    if not businesses:
        print("\n[INFO] No businesses found. Try a different category or city.")
        return

    # Phase 2: Scrape each website
    print("\n" + "-" * 40)
    print("  STEP 2 of 2: Website Scraping")
    print("-" * 40)
    businesses = asyncio.run(scrape_all_websites(businesses))

    # Phase 3: Dedup against DB — remove businesses whose email we already have
    print("\n[DB] Checking for duplicate emails...")
    new_businesses, skipped_count = filter_new_businesses(businesses)
    if skipped_count:
        print(f"[DB] Skipped {skipped_count} businesses already in DB")
    print(f"[DB] {len(new_businesses)} new leads to save")

    # Save new leads (with email) to DB
    saved_count = save_businesses(new_businesses)
    print(f"[DB] Saved {saved_count} new emails to database")

    # Preview and export only the new results
    _print_results_preview(new_businesses)

    print("[EXPORT] Saving results to Excel...")
    filepath = export_to_excel(new_businesses, category, city)
    print(f"[SAVED] Export complete: {filepath}")
    logger.info("Export saved to %s", filepath)

    master_path = update_master_excel(new_businesses)
    print(f"[MASTER] Master file updated: {master_path}")

    # Final summary
    total_scraped = len(businesses)
    total_new = len(new_businesses)
    priority = sum(1 for b in new_businesses if b.is_high_priority())
    no_website_count = sum(1 for b in new_businesses if not b.has_website)
    with_email = sum(1 for b in new_businesses if b.email)
    social_email_count = sum(1 for b in new_businesses if b.email_source and b.email_source != "website")
    with_wa = sum(1 for b in new_businesses if b.whatsapp)

    print()
    print("=" * 60)
    print("  SCRAPE COMPLETE - Summary")
    print("=" * 60)
    print(f"  Total scraped       : {total_scraped}")
    print(f"  Already in DB       : {skipped_count}  (duplicate emails, skipped)")
    print(f"  New leads           : {total_new}")
    print(f"  No website (leads)  : {no_website_count}")
    print(f"  High priority leads : {priority}  (no website OR score < 50)")
    print(f"  With email          : {with_email}  ({social_email_count} found via social media)")
    print(f"  With WhatsApp       : {with_wa}")
    print(f"  Excel saved to      : {filepath}")
    print("=" * 60)


if __name__ == "__main__":
    main()
